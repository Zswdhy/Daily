import re
import time

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
name = "Hypers"
path = f"{name}_{time.strftime('%Y%m%d_%H%M%S', time.localtime())}.xlsx"

ws1 = wb.active
ws1.title = "Profiling纬度-三方标签"

ws1.append(["基本属性 Demographic", "", "人数", "占比"])
ws1.append(["性别", "男", 12, 12])
ws1.append(["", "女", 12, 12])

ws1.append(["年龄", "18-24岁", 12, 12])
ws1.append(["", "25-34岁", 12, 12])
ws1.append(["", "35-44岁", 12, 12])
ws1.append(["", "45岁+", 12, 12])
ws1.append(["", "Mean", 12, 12])

ws1.append(["消费水平", "低", 12, 12])
ws1.append(["", "中", 12, 12])
ws1.append(["", "高", 12, 12])

ws1.append(["职业", "大学生", 12, 12])
ws1.append(["", "白领", 12, 12])
ws1.append(["", "教师", 12, 12])
ws1.append(["", "程序员", 12, 12])
ws1.append(["", "医生", 12, 12])
ws1.append(["", "货车司机", 12, 12])
ws1.append(["", "网约车司机", 12, 12])

ws1.append(["婚姻状况", "单身交友", 12, 12])
ws1.append(["", "未婚", 12, 12])
ws1.append(["", "已婚", 12, 12])

ws1.append(["育儿状态", "备孕", 12, 12])
ws1.append(["", "孕期", 12, 12])
ws1.append(["", "0-1岁小孩父母", 12, 12])
ws1.append(["", "1-3岁小孩父母", 12, 12])
ws1.append(["", "3-6岁小孩父母", 12, 12])
ws1.append(["", "小学生家长", 12, 12])
ws1.append(["", "中学生家长", 12, 12])

ws1.append(["居住城市级别", "一线城市", 12, 12])
ws1.append(["", "二线城市", 12, 12])
ws1.append(["", "三线城市", 12, 12])
ws1.append(["", "四线及以下城市", 12, 12])

ws1.append(["居住Top10城市", "city1", 12, 12])
ws1.append(["", "city2", 12, 12])
ws1.append(["", "city3", 12, 12])
ws1.append(["", "city4", 12, 12])
ws1.append(["", "city5", 12, 12])
ws1.append(["", "city6", 12, 12])
ws1.append(["", "city7", 12, 12])
ws1.append(["", "city8", 12, 12])
ws1.append(["", "city9", 12, 12])
ws1.append(["", "city10", 12, 12])

ws1.append([])
ws1.append([])

ws1.append(['应用分析', "", "", "", ""])
ws1.append(["APP名称", "APP一级标签", "APP二级标签", "人数", "占比"])
ws1.append(["App1", "", "", "", ""])
ws1.append(["App2", "", "", "", ""])
ws1.append(["App3", "", "", "", ""])
ws1.append(["App4", "", "", "", ""])

ws1.append([])
ws1.append([])

ws1.append(['兴趣偏好Interests', "", "", "", ""])
ws1.append(["一级标签", "二级标签", "三级标签", "人数", "占比"])
ws1.append(["Interests1", "", "", "", ""])
ws1.append(["Interests2", "", "", "", ""])
ws1.append(["Interests3", "", "", "", ""])
ws1.append(["Interests4", "", "", "", ""])

# 设置边框样式
side = Side(border_style='thin', color='FF000000')
# 设置字体样式 加粗
font = Font(bold=True)
for item in ["A", "B", "C", "D"]:
    ws1[f"{item}1"].font = font
# 单元格宽度
ws1.column_dimensions["A"].width = 26
ws1.column_dimensions["B"].width = 26
ws1.column_dimensions["C"].width = 26

for rows in ws1.rows:
    if rows[0].value in ["应用分析", "兴趣偏好Interests"]:
        index = re.search(r'\d+>', str(rows[0]))[0][:-1]

        ws1[f"A{index}"].font = font
        cur_row = int(index)
        after_row = int(index) + 1
        for item in ["A", "B", "C", "D", "E"]:
            ws1[f"{item}{after_row}"].font = font
            ws1[f"{item}{cur_row}"].border = Border(top=side, right=side, bottom=side, left=side)

        # 居中显示
        for item in ["A", "B", "C", ]:
            ws1[f"{item}{index}"].alignment = Alignment(horizontal='center', vertical='center')
    for item in rows:
        if item.value is not None:
            ws1[f"{get_column_letter(item.column)}{item.row}"].border = Border(top=side, right=side, bottom=side,
                                                                               left=side)

ws2 = wb.create_sheet(title="Profiling纬度-自有标签")

ws2.append(["", "", "人数", "占比"])
ws2.append(["会员基本信息", "", "", ""])
ws2.append(["会员类别", "潜在会员", "", ""])
ws2.append(["", "R12活跃会员", "", ""])
ws2.append(["", "R24流失会员", "", ""])
ws2.append(["", "R24+流失会员", "", ""])

ws2.append(["活跃会员类型", "R12新会员", "", ""])
ws2.append(["", "R12老会员", "", ""])

ws2.append(["活跃会员等级", "普卡会员", "", ""])
ws2.append(["", "银卡会员", "", ""])
ws2.append(["", "金卡会员", "", ""])
ws2.append(["", "白金卡会员", "", ""])

ws2.append([])
ws2.append([])

ws2.append(["会员消费结构", "", "人数", "占比"])
ws2.append(["活跃会员年消费金额", "9999+", "", ""])
ws2.append(["", "6000-9999", "", ""])
ws2.append(["", "4000-5999", "", ""])
ws2.append(["", "1500-3999", "", ""])
ws2.append(["", "<1500", "", ""])

ws2.append(["活跃会员年消费频次", "0", "", ""])
ws2.append(["", "1", "", ""])
ws2.append(["", "2", "", ""])
ws2.append(["", "3", "", ""])
ws2.append(["", "4", "", ""])
ws2.append(["", "5", "", ""])
ws2.append(["", "6+", "", ""])

ws2.append(["活跃会员年消费件数", "0", "", ""])
ws2.append(["", "1", "", ""])
ws2.append(["", "2", "", ""])
ws2.append(["", "3", "", ""])
ws2.append(["", "4", "", ""])
ws2.append(["", "5", "", ""])
ws2.append(["", "6+", "", ""])

ws2.append([])
ws2.append([])

ws2.append(["CAM KPI", "", "绝对值"])
ws2.append(["", "AUS", ""])
ws2.append(["", "IPT", ""])
ws2.append(["", "Purchase Frequency", ""])
ws2.append(["", "Avg. Spending", ""])
ws2.append(["", "Trasaction", ""])
ws2.append(["", "Qty", ""])
ws2.append(["", "Revenue", ""])
ws2.append(["", "Active Rate", ""])
ws2.append(["", "New Consumer Rate", ""])
ws2.append(["", "New Consumer 2nd Purchase Rate", ""])
ws2.append(["", "Retention Rate", ""])

ws2.append([])
ws2.append([])

ws2.append(["会员沟通触达", "", "绝对值"])
ws2.append(["", "关注公众好比例", ""])
ws2.append(["", "添加企业微信好友比例", ""])
ws2.append(["", "加入企微社群比例", ""])

ws2.append(["会员沟通相应", "", "绝对值"])
ws2.append(["生日礼", "参与/领取率", ""])
ws2.append(["", "领取购买率", ""])
ws2.append(["NCC", "参与/领取率", ""])
ws2.append(["", "领取购买率", ""])
ws2.append(["保级礼", "参与/领取率", ""])
ws2.append(["", "领取购买率", ""])
ws2.append(["升级礼", "参与/领取率", ""])
ws2.append(["", "领取购买率", ""])
ws2.append(["预流失", "参与/领取率", ""])
ws2.append(["", "领取购买率", ""])

ws2.append([])
ws2.append([])

ws2.append(["会员消费偏好", "", "人数"])
ws2.append(["购买产品系列", "系列1", ""])
ws2.append(["", "系列2", ""])
ws2.append(["", "系列N", ""])
ws2.append(["购买产品品类", "品类1", ""])
ws2.append(["", "品类2", ""])
ws2.append(["", "品类N", ""])

ws2.column_dimensions["A"].width = 26
ws2.column_dimensions["B"].width = 26
ws2.column_dimensions["C"].width = 26


def ws_style(ws, column_ls):
    """
    处理 ws 格式
    :param ws:指定的 ws
    :param column_ls:需要处理的列
    :return:
    """
    cur_row = re.search(r'\d+>', str(rows[0]))[0][:-1]
    ws[f"A{index}"].font = font
    cur_row = int(cur_row)
    for item in column_ls:
        ws[f"{item}{cur_row}"].font = font
        ws[f"{item}{cur_row}"].border = Border(top=side, right=side, bottom=side, left=side)


add_filed_ls = ["金额取值参考标准：", "L5  （Top 5%）", "L4  （5% -15%）", "L3  （15% -40%）", "L2  （40% -70%）", "L1 （Bottom 40%）"]

for rows in ws2.rows:
    if str(rows[0].value).strip() in ["会员基本信息", "会员消费结构"]:
        ws_style(ws2, ["A", "B", "C", "D"])
        # ws2 特殊处理的数据列
        if str(rows[0].value).strip() == "会员消费结构":
            cur_row = re.search(r'\d+>', str(rows[0]))[0][:-1]
            for i in range(len(add_filed_ls)):
                ws2[f"F{int(cur_row) + i}"] = add_filed_ls[i]

    if str(rows[0].value).strip() in ["CAM KPI", "会员沟通触达", "会员沟通相应", "会员消费偏好"]:
        ws_style(ws2, ["A", "B", "C"])

    for item in rows:
        if item.value is not None:
            ws2[f"{get_column_letter(item.column)}{item.row}"].border = Border(top=side, right=side, bottom=side,
                                                                               left=side)

ws2["C1"].font = font
ws2["D1"].font = font

ws3 = wb.create_sheet(title="Summary")
ws3.append(["官网首单转化", "", "", "", "", "", "", ""])
ws3.append(["官网首单转化", "", "", "", "", "", "", ""])
ws3.append(["注册场景", "注册人数", "YTD", "YTD", "YTD", "MTD", "MTD", "MTD"])
ws3.append(["注册场景", "注册人数", "转化人数", "转化金额", "转化率", "转化人数", "转化金额", "转化率"])
ws3.append(["朋友圈", "", "", "", "", "", "", ""])
ws3.append(["抖音", "", "", "", "", "", "", ""])
ws3.append(["小红书", "", "", "", "", "", "", ""])

ws3.column_dimensions["A"].width = 26
ws3.merge_cells("A1:H2")  # 左上角到右下角的坐标
ws3["A1"].alignment = Alignment(horizontal='left', vertical='bottom')  # 合并之后，左上角的位置坐标
# ws3["A1"].fill = PatternFill('solid', fgColor="ff00ff")  # 单元格背景颜色
ws3["A1"].font = Font(u'微软雅黑', bold=True, italic=False, strike=False, color="ff0000")  # 文字颜色

ws3.merge_cells("A3:A4")
ws3["A3"].alignment = Alignment(horizontal='center', vertical='center')

ws3.merge_cells("B3:B4")
ws3["B3"].alignment = Alignment(horizontal='center', vertical='center')

ws3.merge_cells("C3:E3")
ws3["C3"].alignment = Alignment(horizontal='center', vertical='center')

ws3.merge_cells("F3:H3")
ws3["F3"].alignment = Alignment(horizontal='center', vertical='center')

wb.save(filename=f'{path}')
