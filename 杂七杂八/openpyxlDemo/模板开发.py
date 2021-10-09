import re
import time

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
name = "Hypers"
path = f"{name}_{time.strftime('%Y%m%d_%H%M%S', time.localtime())}.xlsx"

ws1 = wb.active
ws1.title = "Summary"

ws1.append(["官网首单转化", "1231231231", "132123123", "1231231", "123123", "0.56446464", "0.4641316414341", ""])
ws1.append(["官网首单转化", "", "", "", "", "", "", ""])
ws1.append(["注册场景", "注册人数", "YTD", "YTD", "YTD", "MTD", "MTD", "MTD"])
ws1.append(["注册场景", "注册人数", "转化人数", "转化金额", "转化率", "转化人数", "转化金额", "转化率"])
ws1.append(["朋友圈", "", "", "", "", "", "", ""])
ws1.append(["抖音", "", "", "", "", "", "", ""])
ws1.append(["小红书", "", "", "", "", "", "", ""])

ws1["F1"].number_format = '0.00%'
ws1["G1"].number_format = '0.00%'

# ws1.column_dimensions["A"].width = 26
# ws1.merge_cells("A1:H2")  # 左上角到右下角的坐标
# ws1["A1"].alignment = Alignment(horizontal='left', vertical='bottom')  # 合并之后，左上角的位置坐标
# # ws1["A1"].fill = PatternFill('solid', fgColor="ff00ff")  # 单元格背景颜色
# ws1["A1"].font = Font(u'微软雅黑', bold=True, italic=False, strike=False, color="ff0000")  # 文字颜色
#
# ws1.merge_cells("A3:A4")
# ws1["A3"].alignment = Alignment(horizontal='center', vertical='center')
#
# ws1.merge_cells("B3:B4")
# ws1["B3"].alignment = Alignment(horizontal='center', vertical='center')
#
# ws1.merge_cells("C3:E3")
# ws1["C3"].alignment = Alignment(horizontal='center', vertical='center')
#
# ws1.merge_cells("F3:H3")
# ws1["F3"].alignment = Alignment(horizontal='center', vertical='center')
#
# font = Font(bold=True)
# for item in ["A", "B", "C", "D"]:
#     ws1[f"{item}1"].font = font
wb.save(filename=f'{path}')
